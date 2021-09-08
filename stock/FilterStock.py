#!/usr/bin/python
# -*- coding: utf-8 -*-

# 输入可指定SH/SZ/SH SZ(不指定默认为两个市场)
# 从新华财经获取股票代码
# 逐个代码从雪球获取现金流量表(cash_flow)、利润表(income)、主要指标(indicator)
# 将ReportInfo勒种的四个指标字段输出到output_stock.xlsx文件中
import sys
import time
import os
from CommonRequest import *

try:
    import xml.etree.cElementTree as ET
except ImportError:
    import xml.etree.ElementTree as ET

from StockInfo import *
import openpyxl
from openpyxl.styles import Alignment, Font, NamedStyle

comm_url = "https://stock.xueqiu.com/v5/stock/finance/cn/{}.json?symbol={}&type=Q4&is_detail=true&count=5&timestamp="
market_stock_url = "https://market2.cnfic.com.cn/quote/v1/market/detail?finance_mic={}"


def get_request(url, stock, field, field_bak, func):
    while True:
        url += str(int(round(time.time()*1000)))
        jsondata = XueqiuRequest().get_data(url).json()
        if "data" not in jsondata:
            print("request {} fail. response content: {}".format(url, jsondata))
            val = random.randint(1, 5)
            time.sleep(val)
            continue
        break

    datalist = jsondata["data"]["list"]
    for ele in datalist:
        name = ele["report_name"]
        value = ele.get(field_bak) if ele.get(field) is None else ele.get(field)
        if value is None:
            print("({},{}) table{} field{} field-bak{} fail".format(stock.id, stock.symbol, url, field, field_bak))
            continue
        if len(value) > 0:
            func(stock, name, value[0])


def get_data(code, stock):
    get_request(comm_url.format("cash_flow", code), stock, "cash_received_of_sales_service", "cash_received_of_interest_etc", StockInfo.add_cashflow)
    get_request(comm_url.format("income", code), stock, "revenue", "total_revenue", StockInfo.add_revenue)
    get_request(comm_url.format("indicator", code), stock, "avg_roe", "", StockInfo.add_roe)


def writeSheetHead(stock, sheet):
    sheet.merge_cells('A1:A2')
    sheet['A1'].value = "代码"
    sheet['A1'].style = "style_titleRow"

    sheet.merge_cells('B1:B2')
    sheet['B1'].value = "名称"
    sheet['B1'].style = "style_titleRow"

    start = 3
    step = 3
    for reportname in stock.reports.keys():
        sheet.merge_cells(start_row=1, end_row=1, start_column=start, end_column=start + step)
        sheet.cell(1, start).value = reportname
        sheet.cell(1, start).style = "style_titleRow"

        sheet.cell(2, start).value = "销售商品、提供劳务收到的现金"
        sheet.cell(2, start).style = "style_titleRow"

        sheet.cell(2, start + 1).value = "营业总收入"
        sheet.cell(2, start + 1).style = "style_titleRow"

        sheet.cell(2, start + 2).value = "现金收入比"
        sheet.cell(2, start + 2).style = "style_titleRow"

        sheet.cell(2, start + 3).value = "净资产收益率"
        sheet.cell(2, start + 3).style = "style_titleRow"

        start += step + 1


def get_value(val):
    if val is None:
        return "--"
    if isinstance(val, int):
        return str(val)
    if isinstance(val, float):
        val1 = str(val).rstrip('0')
        return val1.rstrip('.') if val1.endswith('.') else str(val)


def write_excel(stocks, filename, sheetname):
    if len(stocks) <= 0:
        return

    if os.path.exists(filename):
        workbook = openpyxl.load_workbook(filename)
    else:
        workbook = openpyxl.Workbook()
        workbook.add_named_style(NamedStyle(name="style_titleRow", font=Font(bold=True),
                   alignment=Alignment(horizontal="center", vertical="center")))

    sheet = workbook.create_sheet(sheetname, 0)
    writeSheetHead(stocks[0], sheet)

    row = 3
    for stock in stocks:
        sheet.cell(row, 1, stock.id)
        sheet.cell(row, 2, stock.symbol)
        col = 3
        for report in list(stock.reports.values()):
            sheet.cell(row, col, get_value(report.cashflow))
            col += 1
            sheet.cell(row, col, get_value(report.revenue))
            col += 1
            sheet.cell(row, col, get_value(report.cash_revenue_ratio))
            col += 1
            sheet.cell(row, col, get_value(report.roe))
            col += 1
        row += 1
    workbook.save(filename)


def get_all_stocks(mkt):
    market_map = {"SH":"XSHG.ESA.M", "SZ": "XSHE.ESA.M"}
    url = market_stock_url.format(market_map.get(mkt))
    response_data = CommonRequest().get_data(url).json()
    if "data" not in response_data:
        print("request {} fail. response content: {}".format(url, response_data))
        return
    market_detail_prod_grp = response_data["data"]["market_detail_prod_grp"]
    code_list = []
    for ele in market_detail_prod_grp:
        code_list.append({"code":ele["prod_code"],"name":ele["prod_name"]})
    return code_list


if __name__ == "__main__":
    market_list = []
    if len(sys.argv) > 1:
        market_list = sys.argv[1:]
    else:
        market_list = ['SH', 'SZ']

    xlsx_filename = "output_stock.xlsx"
    for market in market_list:
        codes = get_all_stocks(market)
        stocks = []
        for code_info in codes:
            stock = StockInfo(code_info["code"], code_info["name"])
            get_data(market + code_info["code"], stock)
            stocks.append(stock)
            print("{},{}".format(code_info["code"], code_info["name"]))
        write_excel(stocks, xlsx_filename, market)
